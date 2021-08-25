import openpyxl.cell.text
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import Reference, LineChart
from openpyxl.styles import Font, Fill, Color, PatternFill, Border
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font
from openpyxl.styles.borders import Border, Side


# Pandas to read the Excel file, create a pivot table, and export it to Excel
# Then we’ll use the Openpyxl library to write Excel formulas, make charts and format the spreadsheet through Python


excel_file = pd.read_excel('Impulse Noise Evaluation.xlsx')
var = excel_file[[]]
report_table = excel_file.pivot_table(index='Measurement time')
report_table.to_excel('report_2021.xlsx', sheet_name='Report', startrow=0)  # Creates report in excel

wb = load_workbook('report_2021.xlsx')  # variable to load workbook
sheet = wb['Report']  # assign sheet

# Font

grayFill = PatternFill(start_color='C0C0C0',
                       end_color='C0C0C0',
                       fill_type='solid')


# Autofit columns
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

#  Format
#  property cell.border should be used instead of cell.style.border

thin = Side(border_style="thin", color="000000")  # Border style, color
border = Border(left=thin, right=thin, top=thin, bottom=thin)  # Position of border
for row in sheet['A1:W4']:
    for cell in row:
        cell.border = border  # A5:W5 range cell setting border

for col in sheet.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name

    for cell in col:
        if cell.value:  # Fills cell color in range
            sheet[cell.coordinate].fill = grayFill

        if cell.coordinate in sheet.merged_cells:  # not check merge_cells

            continue
        try:  # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)

        except:
            pass
    adjusted_width = (max_length + 2) * 1.2

    sheet.column_dimensions[column].width = adjusted_width

# cell references (original spreadsheet)
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# barchart
linechart = LineChart()
# locate data and categories
data = Reference(sheet,
                 min_col=min_column + 1,
                 max_col=max_column,
                 min_row=min_row,
                 max_row=max_row)
# including headers
categories = Reference(sheet,
                       min_col=min_column,
                       max_col=min_column,
                       min_row=min_row + 1,
                       max_row=max_row)  # not including headers

# adding data and categories
linechart.add_data(data, titles_from_data=True)
linechart.set_categories(categories)
openpyxl.styles.alignment.Alignment(horizontal='center')

# Format chart
font_chart = Font(typeface='Calibri')
cp = CharacterProperties(latin=font_chart, sz=1500)

linechart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

# line style


# location chart

sheet.add_chart(linechart, "B12")
linechart.height = 13.0
linechart.width = 25.0
linechart.title = 'Impulse Noise Evaluation'
linechart.style = 2  # choose the chart style
linechart.x_axis.title = "Test Shot"
wb.save('report_2021.xlsx')

print(report_table)
# print(excel_file)
